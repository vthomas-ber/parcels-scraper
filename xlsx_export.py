"""
xlsx_export.py — styled DMF-ready Excel export for Food Data Researcher PRO.

Produces a two-sheet workbook:
  1. "DMF Upload"  — color-coded clusters, frozen header, audit columns hidden
                     (grouped, so users can expand them with the +/- outline)
  2. "Legend"      — color key, status definitions, reliability tiers

Usage in app.py (after edited_df exists):

    from xlsx_export import build_dmf_workbook
    xlsx_bytes = build_dmf_workbook(edited_df.drop(columns=["Re-run?"], errors="ignore"))
    st.download_button(
        "⬇️ Download DMF-ready Excel",
        data=xlsx_bytes,
        file_name=f"dmf_export_{market_code}_{time.strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
"""

from io import BytesIO
import base64
import json

import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Cluster definitions ──────────────────────────────────────────────────────
# name -> (header fill, body fill, column list, hidden-by-default)
CLUSTERS = {
    "Identity": {
        "header": "9DC3E6", "body": "DDEBF7", "hidden": False,
        "cols": ["Product Image", "GTIN / EAN", "User Input", "Product Name", "Brand"],
    },
    "Validation": {
        "header": "F4B183", "body": "FBE5D6", "hidden": False,
        "cols": ["Status", "Info Reliability"],
    },
    "Product category": {
        "header": "FFD966", "body": "FFF2CC", "hidden": False,
        "cols": ["Category L1", "Category L2", "Category L3",
                 "Category L4", "Category L5", "Category L6"],
    },
    "Product attributes": {
        "header": "F8CBAD", "body": "FCE4D6", "hidden": False,
        "cols": ["UoM", "Fragile Item", "Net Weight (g) / Volume",
                 "Gross Weight (g)", "Organic Product",
                 "Net Weight/ Volume (Customer Facing)",
                 "Dietary Tags", "Occasion Tags", "Seasonal Tags",
                 "Organic Certification ID"],
    },
    "Food information": {
        "header": "A9D08E", "body": "E2EFDA", "hidden": False,
        "cols": ["Ingredients", "Allergens", "May Contain", "Nutritional Info",
                 "Manufacturer Name", "Manufacturer Address", "Place of Origin",
                 "Energy (kJ)", "Fat (g)", "Of Which Saturated Fatty Acids (g)",
                 "Carbohydrates (g)", "Of Which Sugars (g)", "Protein (g)",
                 "Fiber (g)", "Salt (g)"],
    },
    "Source / link metadata (audit)": {
        "header": "BFBFBF", "body": "D9D9D9", "hidden": True,
        "cols": ["Image 1", "Image 2", "Image Source Link",
                 "Source 1", "Source 2", "Source 3", "Source 4", "Source 5",
                 "Source Candidates (audit)"],
    },
    "Internal rationale (audit)": {
        "header": "D0CECE", "body": "EDEDED", "hidden": True,
        "cols": ["Reliability Reasoning", "Chain of Thought",
                 "Categorization Diagnosis", "Tagging Reasoning",
                 "Image 2 Failure Reason", "Cached"],
    },
}

# Per-cell conditional fills for validation values
STATUS_FILLS = {
    "Success":            "C6EFCE",   # green
    "⚠️ Needs Review":    "FFEB9C",   # amber
    "Failed Validation":  "FFC7CE",   # red
}
RELIABILITY_FILLS = {"H": "C6EFCE", "M": "FFEB9C", "L": "FFC7CE"}

_THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def _cluster_of(col: str):
    for name, spec in CLUSTERS.items():
        if col in spec["cols"]:
            return name, spec
    return None, None


def build_dmf_workbook(df: pd.DataFrame) -> bytes:
    wb = Workbook()

    # ── Sheet 1: DMF Upload ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "DMF Upload"

    # Pull the serialised verified image bytes out of the frame (hidden helper
    # column produced by process_ean) and drop it from the visible sheet.
    embedded_by_rowidx = {}
    if "_EmbeddedImages" in df.columns:
        for pos, raw in enumerate(df["_EmbeddedImages"].tolist()):
            if isinstance(raw, str) and raw.strip():
                try:
                    embedded_by_rowidx[pos] = json.loads(raw)
                except Exception:
                    embedded_by_rowidx[pos] = []
        df = df.drop(columns=["_EmbeddedImages"])

    # Ensure a Product Image column exists (holds the embedded picture; the
    # cell text stays blank). Only add it if we actually have images to embed.
    has_embedded = any(embedded_by_rowidx.values())
    if has_embedded and "Product Image" not in df.columns:
        df.insert(0, "Product Image", "")

    # Column order: visible clusters first (in CLUSTERS order), hidden last,
    # then any column not mapped to a cluster (kept visible so nothing is lost).
    ordered = []
    for spec in CLUSTERS.values():
        ordered += [c for c in spec["cols"] if c in df.columns]
    ordered += [c for c in df.columns if c not in ordered]
    df = df[ordered]

    img_col_idx = (df.columns.get_loc("Product Image") + 1) if "Product Image" in df.columns else None

    # Header row
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        _, spec = _cluster_of(col)
        fill = (spec or {}).get("header", "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = _THIN

    # Body
    for i, (pos, (_, row)) in enumerate(zip(range(len(df)), df.iterrows()), start=2):
        row_has_image = bool(embedded_by_rowidx.get(pos))
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value="" if pd.isna(val) else val)
            _, spec = _cluster_of(col)
            body_fill = (spec or {}).get("body")
            # Validation cells get value-driven colors overriding cluster fill
            if col == "Status":
                body_fill = STATUS_FILLS.get(str(val).strip(), body_fill)
            elif col == "Info Reliability":
                body_fill = RELIABILITY_FILLS.get(str(val).strip(), body_fill)
            if body_fill:
                cell.fill = PatternFill("solid", fgColor=body_fill)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=(col in (
                "Ingredients", "Nutritional Info", "Chain of Thought",
                "Reliability Reasoning", "Categorization Diagnosis",
                "Tagging Reasoning")))
            # Real hyperlinks for link columns (survive Excel, unlike pasted text)
            if col.startswith("Source ") or col in ("Image Source Link",
                                                    "Image 1", "Image 2"):
                v = str(val or "")
                if v.startswith("http"):
                    cell.hyperlink = v
                    cell.font = Font(color="0563C1", underline="single", size=9)

        # ── Embed the verified product image directly into the cell ──────────
        # The image travels INSIDE the workbook — it cannot 404 or 403 for the
        # next viewer, unlike an exported retailer/CDN URL. Bytes were both
        # downloaded and vision-verified by the pipeline this run.
        if img_col_idx and row_has_image:
            try:
                first = embedded_by_rowidx[pos][0]
                bio = BytesIO(base64.b64decode(first["b64"]))
                xlimg = XLImage(bio)
                # Scale to ~90px tall, preserve aspect ratio
                if xlimg.height:
                    scale = 90 / xlimg.height
                    xlimg.height = 90
                    xlimg.width = int(xlimg.width * scale)
                ws.add_image(xlimg, f"{get_column_letter(img_col_idx)}{i}")
                ws.row_dimensions[i].height = 70
            except Exception:
                ws.cell(row=i, column=img_col_idx, value="(image unavailable)")

    # Column widths + hide audit clusters (grouped so users can expand)
    for j, col in enumerate(df.columns, start=1):
        letter = get_column_letter(j)
        if col == "Product Image":
            ws.column_dimensions[letter].width = 16
        else:
            ws.column_dimensions[letter].width = 14 if len(col) < 16 else 22
        _, spec = _cluster_of(col)
        if spec and spec["hidden"]:
            ws.column_dimensions[letter].hidden = True
            ws.column_dimensions[letter].outlineLevel = 1
    ws.sheet_properties.outlinePr.summaryRight = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # ── Sheet 2: Legend ──────────────────────────────────────────────────────
    lg = wb.create_sheet("Legend")
    lg.column_dimensions["A"].width = 34
    lg.column_dimensions["B"].width = 90

    def write(row, a, b="", a_fill=None, bold=False):
        ca = lg.cell(row=row, column=1, value=a)
        cb = lg.cell(row=row, column=2, value=b)
        if a_fill:
            ca.fill = PatternFill("solid", fgColor=a_fill)
        if bold:
            ca.font = Font(bold=True, size=11)
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        return row + 1

    r = write(1, "LEGEND — how to read this file", bold=True)
    r = write(r, "", "")
    r = write(r, "Color groups", bold=True)
    for name, spec in CLUSTERS.items():
        desc = {
            "Identity": "Barcode, your original input line, and the resolved product name/brand.",
            "Validation": "Overall row status and the reliability grade of the food information.",
            "Product category": "Six-level category assignment against the internal taxonomy.",
            "Product attributes": "Logistics and merchandising attributes plus dietary/occasion/seasonal tags.",
            "Food information": "Legally relevant food data: ingredients, allergens, nutrition per 100g/ml, manufacturer.",
            "Source / link metadata (audit)": "HIDDEN by default. Image and source URLs kept for audit. Unhide via the + outline button or Format → Unhide Columns.",
            "Internal rationale (audit)": "HIDDEN by default. AI reasoning trails explaining reliability, categorisation, and tagging decisions.",
        }.get(name, "")
        r = write(r, name, desc, a_fill=spec["body"])

    r = write(r + 1, "Status definitions", bold=True)
    r = write(r, "Success", "Data found, validated against the EAN, and reliability H (or M with barcode verification). Safe for DMF upload.", a_fill="C6EFCE")
    r = write(r, "⚠️ Needs Review", "Data found but either reliability is Low, no approved retail source confirmed it, or the barcode could not be independently verified. Human check required before upload.", a_fill="FFEB9C")
    r = write(r, "Failed Validation", "The product found online does NOT match the EAN/your input. All food data suppressed. Do not upload.", a_fill="FFC7CE")
    r = write(r, "⚠️ Not Found", "No approved source indexed this EAN. Consider re-running with a product name added to the input line.", a_fill="D9D9D9")

    r = write(r + 1, "Info Reliability tiers", bold=True)
    r = write(r, "H (High)", "Data cross-confirmed by a Tier-1 source (barcode registry or approved retailer page containing the EAN).", a_fill="C6EFCE")
    r = write(r, "M (Medium)", "Data from a real retail/brand source but without independent barcode confirmation.", a_fill="FFEB9C")
    r = write(r, "L (Low)", "Single weak source or inferred data. Always routed to Needs Review.", a_fill="FFC7CE")

    r = write(r + 1, "Source tiers (audit columns)", bold=True)
    r = write(r, "Tier 1", "Barcode-keyed registries: Go-UPC, EAN-Search.org (barcode match verified).")
    r = write(r, "Tier 2", "Approved retailer/brand pages where the EAN was confirmed in the page itself.")
    r = write(r, "Tier 3", "Approved retailer pages found via search but without on-page EAN confirmation.")
    r = write(r, "Excluded", "Amazon, eBay, AliExpress, OpenFoodFacts — never used for data or images.")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
